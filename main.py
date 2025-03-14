import streamlit as st
import pandas as pd
import io
import zipfile
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from copy import copy
import numpy as np


# set page configurations
st.set_page_config(
    page_title="Cleaning App",
    layout="centered",
    initial_sidebar_state="collapsed",
    page_icon=":material/mop:"
)

# the custom CSS lives here:
hide_default_format = """
    <style>
        .reportview-container .main footer {visibility: hidden;}    
        #MainMenu, header, footer {visibility: hidden;}
        div.stActionButton{visibility: hidden;}
        [class="stAppDeployButton"] {
            display: none;
        }
        [data-testid="collapsedControl"] {
            display: none
        }
        [data-testid="stFileUploaderDropzone"] {
            background-color: #9cdcf3;
            border-radius: 15px;
            border: 1px solid #1F2041;
            padding: 15px; /* Optional: Add some padding */
        }
        div[data-testid="stFileUploaderDropzoneInstructions"]>div>span {
            visibility: hidden;
        }
        div[data-testid="stFileUploaderDropzoneInstructions"]>div>span::before {
            content: "Drag & drop Survey Monkey results for cleaning.";
            visibility: visible;
        }
        div[data-testid="stFileUploaderDropzoneInstructions"]>div>small {
            visibility: hidden;
        }
        div[data-testid="stFileUploaderDropzoneInstructions"]>div>small::before {
            content: "You can upload more than 1 file at a time!";
            visibility: visible;
        }
        .stDownloadButton, div.stButton {text-align:center}
        [data-testid="stBaseButton-secondary"] {
            background-color: #9cdcf3;
            border-radius: 15px;
            border: 1px solid #1F2041;
        }
        [data-testid="stFileUploaderFile"] {
            color: #fefefe;
        }
        div[data-testid="stFileUploaderFile"] div>small {
            color: #fefefe;
        }
        [data-testid="stFileUploaderPagination"] >small {
            color: #fefefe;
        }
        [data-testid="stBaseButton-minimal"] {
            color: #fefefe;
        }
        [data-testid="stImageContainer"] {
            position: absolute;
            top: 0%; 
            left: -210%; 
        }
        [data-testid="stBaseButton-elementToolbar"] {
            display: none;
        }
    </style>
"""

# inject the CSS
st.markdown(hide_default_format, unsafe_allow_html=True)

# heading text
# fefefe (white)
# 9cdcf3 (blue)
# 1b1b1b (black)
st.markdown(f'''
    <p style="font-size: 40px; font-weight: 900; text-align: center; margin-top: 10px; margin-bottom: 80px; color: #9cdcf3;">
        11 Ten Leadership Cleaning App
    </p>
''', unsafe_allow_html=True)


# Define helper function to extract the question number from the string for sorting
def extract_question_number(question):
    try:
        return int(question.split()[0][1:])  # Extract the number following 'Q'
    except ValueError:
        return None  # If no number is found, return None


def main():
    # declare variable for uploading files
    uploaded_files = st.file_uploader(
        label="Choose completed reporting template",
        label_visibility='collapsed',
        accept_multiple_files=True,
        help="Upload Survey Monkey template(s) here."
    )

    if uploaded_files:

        # Prepare list to hold file data
        files_to_download = []

        # capture which templates are uploaded
        templates_list = []

        # Iterate through uploaded files
        for uploaded_file in uploaded_files:

            # check the file extension of the uploaded file. If it's not XLSX, return an error message
            if uploaded_file.name.split('.')[-1] != 'xlsx':
                st.error(
                    "The uploaded file is not in the correct format. Please upload an Excel file.")
                return

            # Read the file into memory
            file_data = io.BytesIO(uploaded_file.read())

            # Load the workbook using openpyxl
            wb = openpyxl.load_workbook(file_data)
            ws = wb.active  # Assuming the data is on the active sheet

            # Iterate over all merged cells and unmerge them
            for merged_cell_range in list(ws.merged_cells.ranges):
                ws.unmerge_cells(str(merged_cell_range))

            # check to see if the uploaded file has already been cleaned by seeing if cell D23 is empty. If it is, we're good to go.
            if ws["D23"].value is not None:
                st.error(
                    "The uploaded file appears to have been processed already! Please upload a different file.")
                return

            # Add new columns & formatting
            ws["D23"] = "Question Order"
            ws["D23"].font = copy(ws["C23"].font)
            ws["D23"].fill = copy(ws["C23"].fill)
            ws["D23"].alignment = copy(ws["C23"].alignment)

            ws["E23"] = "1st-Order Category"
            ws["E23"].font = copy(ws["C23"].font)
            ws["E23"].fill = copy(ws["C23"].fill)
            ws["E23"].alignment = copy(ws["C23"].alignment)

            ws["F23"] = "2nd-Order Category"
            ws["F23"].font = copy(ws["C23"].font)
            ws["F23"].fill = copy(ws["C23"].fill)
            ws["F23"].alignment = copy(ws["C23"].alignment)

            # Extract the data starting from A23 down until the first blank cell
            row = 24
            data = []
            while ws[f"A{row}"].value is not None:
                # Append row data into the list
                data.append(
                    [ws[f"A{row}"].value, ws[f"B{row}"].value, ws[f"C{row}"].value])
                row += 1

            # Create a DataFrame from the extracted data
            df = pd.DataFrame(
                data, columns=["Questions", "Difficulty", "Average Score"])

            # Convert the 'Average Score' column to integer
            df["Average Score"] = df["Average Score"].str.replace(
                '%', '').astype(int)
            df = df.rename(columns={"Average Score": "Avg. Score (%)"})

            # convert overall average
            overall_avg = df["Avg. Score (%)"].mean()

            # Add the "Question Order" column based on the extracted question
            df["Question Order"] = df["Questions"].apply(
                extract_question_number).astype(int)

            # Sort the DataFrame by "Question Order"
            df = df.sort_values(by="Question Order")

            # Overwrite the "Question Order" column with sequential numbers starting from 1
            df["Question Order"] = range(1, len(df) + 1)

            # 1st-, 2nd-, and 3rd-Order Categories
            category_1 = ['THRIVE', 'Just Leader']
            category_2 = [
                'Trust', 'Health', 'Relationships', 'Impact', 'Value', 'Engagement', 'See the Whole Playing Field', 'Build Cultural Competency', 'Give Power Away', 'Take Bold, Courageous Action']

            # this level will only pertain to the LEADER and TEAM templates
            category_3 = [
                'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team', 'Leader', 'Team']

            template = ''

            # fill out templates
            if df.shape[0] == 38:  # Review template

                template = 'Review'
                templates_list.append(template)

                # fill in the "Category" column
                repetitions_1 = [30, 8]
                category_list_1 = []
                for category, rep in zip(category_1, repetitions_1):
                    category_list_1.extend([category] * rep)
                df['1st-Order Category'] = category_list_1
                first0_summary = df.groupby(
                    '1st-Order Category')['Avg. Score (%)'].mean().reset_index()

                # Start from A22 and find the last filled cell in column A
                current_row = 22
                while ws[f"A{current_row}"].value is not None:  # Check if cell is filled
                    current_row += 1  # Move down

                # Move one row below the last filled cell
                start_row = current_row + 1

                # Insert blank rows after the last row written
                ws.insert_rows(current_row, 8)

                # Add the overall average
                ws["A63"] = "Overall Average (%)"
                ws["B63"] = overall_avg

                # use a custom sort for the summary table
                first0_summary['1st-Order Category'] = pd.Categorical(
                    first0_summary['1st-Order Category'], [
                        "THRIVE", "Just Leader",
                    ])
                first0_summary = first0_summary.sort_values(
                    "1st-Order Category")

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                first0_new = first0_summary.set_index("1st-Order Category").T

                # header column
                ws["A65"] = "1st-Order Category"
                ws["A66"] = "Avg. Score (%)"

                # copy in heading labels
                ws["B65"] = "THRIVE"
                ws["C65"] = "Just Leader"

                row = 66  # Starting point
                for _, row_data in first0_new.iterrows():
                    ws[f"B{row}"] = row_data["THRIVE"]
                    ws[f"C{row}"] = row_data["Just Leader"]
                    row += 1  # Move to the next row

                repetitions_2 = [5, 5, 5, 5, 5, 5, 2, 2, 2, 2]
                category_list_2 = []
                for category, rep in zip(category_2, repetitions_2):
                    category_list_2.extend([category] * rep)
                df['2nd-Order Category'] = category_list_2
                second0_summary = df.groupby(
                    ['1st-Order Category', '2nd-Order Category'])['Avg. Score (%)'].mean().reset_index().drop(columns='1st-Order Category')

                # use a custom sort for the summary table
                second0_summary['2nd-Order Category'] = pd.Categorical(
                    second0_summary['2nd-Order Category'], [
                        "Trust", "Health", "Relationships", "Impact", "Value", "Engagement", "See the Whole Playing Field", "Build Cultural Competency", "Give Power Away", "Take Bold, Courageous Action"
                    ])
                second0_summary = second0_summary.sort_values(
                    "2nd-Order Category")

                # header rows
                ws["A68"] = "2nd-Order Category"
                ws["A69"] = "Avg. Score (%)"

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                second0_new = second0_summary.set_index("2nd-Order Category").T

                # New code for wide format
                row = 69  # Starting point
                for _, row_data in second0_new.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Health"]
                    ws[f"D{row}"] = row_data["Relationships"]
                    ws[f"E{row}"] = row_data["Impact"]
                    ws[f"F{row}"] = row_data["Value"]
                    ws[f"G{row}"] = row_data["Engagement"]
                    ws[f"H{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"I{row}"] = row_data["Build Cultural Competency"]
                    ws[f"J{row}"] = row_data["Give Power Away"]
                    ws[f"K{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                # Copy in 2nd-Order Category labels
                ws["B68"] = "Trust"
                ws["C68"] = "Health"
                ws["D68"] = "Relationships"
                ws["E68"] = "Impact"
                ws["F68"] = "Value"
                ws["G68"] = "Engagement"
                ws["H68"] = "See the Whole Playing Field"
                ws["I68"] = "Build Cultural Competency"
                ws["J68"] = "Give Power Away"
                ws["K68"] = "Take Bold, Courageous Action"

                # Format the newly-created cells
                start_row = 24
                end_row = 70
                start_col = 1  # Column A (1-indexed)
                end_col = 11

                # Define the font style
                custom_font = Font(name="Arial", size=11)

                # Apply the font style to each cell in the range
                for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.font = custom_font

                # format the headers - Overall average
                ws["A63"].font = Font(name="Arial", size=12, bold=True)
                ws["B63"].font = Font(name="Arial", size=12, bold=True)

                # First-Order Average
                ws["A65"].font = Font(name="Arial", size=11, bold=True)
                ws["A66"].font = Font(name="Arial", size=11, bold=True)

                # Second-Order Average
                ws["A68"].font = Font(name="Arial", size=11, bold=True)
                ws["A69"].font = Font(name="Arial", size=11, bold=True)

                # Center alignment
                ws["A65"].alignment = copy(ws["C23"].alignment)
                ws["A66"].alignment = copy(ws["C23"].alignment)
                ws["A68"].alignment = copy(ws["C23"].alignment)
                ws["A69"].alignment = copy(ws["C23"].alignment)

                # rounding - summary tables
                for row in ws.iter_rows(min_row=63, max_row=71, min_col=2, max_col=11):
                    for cell in row:
                        # Check if cell contains a numeric value
                        if isinstance(cell.value, (int, float)):
                            # Round to 1 decimal place
                            cell.number_format = '0.0'

                # row height
                for row in range(63, 71):
                    ws.row_dimensions[row].height = 15

                # get standard deviation values in template
                std_dev_values = []

                # Start iterating from the first cell in column C
                column = 3  # Column C
                row = 1  # Start at the first row
                max_row = 5000

                # Iterate through column C until the end and collect STD values
                while row <= max_row:
                    cell_value = ws.cell(row=row, column=column).value

                    # Check if the cell contains 'Standard Deviation'
                    if cell_value == "Standard Deviation":
                        # Get the value in the cell below
                        next_cell_value = ws.cell(
                            row=row + 1, column=column).value
                        # if next_cell_value is not None, append to std_dev_values
                        if next_cell_value is not None:
                            std_dev_values.append(next_cell_value)

                        # Move the row pointer down by 2 to skip the value we just processed
                        row += 2
                    else:
                        # Move to the next row
                        row += 1

            elif df.shape[0] == 47:  # No leader template

                template = 'No leader'
                templates_list.append(template)

                repetitions_1 = [35, 12]
                category_list_1 = []
                for category, rep in zip(category_1, repetitions_1):
                    category_list_1.extend([category] * rep)
                df['1st-Order Category'] = category_list_1
                first0_summary = df.groupby(
                    '1st-Order Category')['Avg. Score (%)'].mean().reset_index()

                # Start from A22 and find the last filled cell in column A
                current_row = 22
                while ws[f"A{current_row}"].value is not None:  # Check if cell is filled
                    current_row += 1  # Move down

                # Move one row below the last filled cell
                start_row = current_row + 1

                # Insert 12 blank rows after the last row written
                ws.insert_rows(current_row, 8)

                # Add the overall average
                ws["A72"] = "Overall Average (%)"
                ws["B72"] = overall_avg

                # use a custom sort for the summary table
                first0_summary['1st-Order Category'] = pd.Categorical(
                    first0_summary['1st-Order Category'], [
                        "THRIVE", "Just Leader",
                    ])
                first0_summary = first0_summary.sort_values(
                    "1st-Order Category")

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                first0_new = first0_summary.set_index("1st-Order Category").T

                # header column
                ws["A74"] = "1st-Order Category"
                ws["A75"] = "Avg. Score (%)"

                # copy in heading labels
                ws["B74"] = "THRIVE"
                ws["C74"] = "Just Leader"

                row = 75  # Starting point
                for _, row_data in first0_new.iterrows():
                    ws[f"B{row}"] = row_data["THRIVE"]
                    ws[f"C{row}"] = row_data["Just Leader"]
                    row += 1  # Move to the next row

                repetitions_2 = [5, 10, 5, 5, 5, 5, 3, 3, 3, 3]
                category_list_2 = []
                for category, rep in zip(category_2, repetitions_2):
                    category_list_2.extend([category] * rep)
                df['2nd-Order Category'] = category_list_2
                second0_summary = df.groupby(
                    ['1st-Order Category', '2nd-Order Category'])['Avg. Score (%)'].mean().reset_index().drop(columns='1st-Order Category')

                # use a custom sort for the summary table
                second0_summary['2nd-Order Category'] = pd.Categorical(
                    second0_summary['2nd-Order Category'], [
                        "Trust", "Health", "Relationships", "Impact", "Value", "Engagement", "See the Whole Playing Field", "Build Cultural Competency", "Give Power Away", "Take Bold, Courageous Action"
                    ])
                second0_summary = second0_summary.sort_values(
                    "2nd-Order Category")

                # header rows
                ws["A77"] = "2nd-Order Category"
                ws["A78"] = "Avg. Score (%)"

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                second0_new = second0_summary.set_index("2nd-Order Category").T

                # New code for wide format
                row = 78  # Starting point
                for _, row_data in second0_new.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Health"]
                    ws[f"D{row}"] = row_data["Relationships"]
                    ws[f"E{row}"] = row_data["Impact"]
                    ws[f"F{row}"] = row_data["Value"]
                    ws[f"G{row}"] = row_data["Engagement"]
                    ws[f"H{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"I{row}"] = row_data["Build Cultural Competency"]
                    ws[f"J{row}"] = row_data["Give Power Away"]
                    ws[f"K{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                # Copy in 2nd-Order Category labels
                ws["B77"] = "Trust"
                ws["C77"] = "Health"
                ws["D77"] = "Relationships"
                ws["E77"] = "Impact"
                ws["F77"] = "Value"
                ws["G77"] = "Engagement"
                ws["H77"] = "See the Whole Playing Field"
                ws["I77"] = "Build Cultural Competency"
                ws["J77"] = "Give Power Away"
                ws["K77"] = "Take Bold, Courageous Action"

                # Define the range of cells
                start_row = 24
                end_row = 79
                start_col = 1  # Column A (1-indexed)
                end_col = 11  # Column C (1-indexed)

                # Define the font style
                custom_font = Font(name="Arial", size=11)

                # Apply the font style to each cell in the range
                for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.font = custom_font

                # format the headers - overall average
                ws["A72"].font = Font(name="Arial", size=12, bold=True)
                ws["B72"].font = Font(name="Arial", size=12, bold=True)

                # First-Order Average
                ws["A74"].font = Font(name="Arial", size=11, bold=True)
                ws["A75"].font = Font(name="Arial", size=11, bold=True)

                # Second-Order Average
                ws["A77"].font = Font(name="Arial", size=11, bold=True)
                ws["A78"].font = Font(name="Arial", size=11, bold=True)

                ws["A74"].alignment = copy(ws["C23"].alignment)
                ws["A75"].alignment = copy(ws["C23"].alignment)
                ws["A77"].alignment = copy(ws["C23"].alignment)
                ws["A78"].alignment = copy(ws["C23"].alignment)

                # rounding
                for row in ws.iter_rows(min_row=72, max_row=79, min_col=2, max_col=11):
                    for cell in row:
                        # Check if cell contains a numeric value
                        if isinstance(cell.value, (int, float)):
                            # Round to 1 decimal place
                            cell.number_format = '0.0'

                # row height
                for row in range(71, 79):
                    ws.row_dimensions[row].height = 15

                # get standard deviation values in template
                std_dev_values = []

                # Start iterating from the first cell in column C
                column = 3  # Column C
                row = 1  # Start at the first row
                max_row = 5000

                # Iterate through column C until the end
                while row <= max_row:
                    cell_value = ws.cell(row=row, column=column).value

                    # Check if the cell contains 'Standard Deviation'
                    if cell_value == "Standard Deviation":
                        # Get the value in the cell below
                        next_cell_value = ws.cell(
                            row=row + 1, column=column).value
                        # if next_cell_value is not None, append to std_dev_values as a float
                        if next_cell_value is not None:
                            next_cell_value = float(next_cell_value)
                            std_dev_values.append(next_cell_value)

                        # Move the row pointer down by 2 to skip the value we just processed
                        row += 2
                    else:
                        # Move to the next row
                        row += 1

            elif str(df["Questions"].iloc[0]).startswith("Q7"):  # Leader template

                template = 'Leader'
                templates_list.append(template)

                df_leader = df.copy()

                # First level breakdown
                repetitions_1 = [60, 20]
                category_list_1 = []
                for category, rep in zip(category_1, repetitions_1):
                    category_list_1.extend([category] * rep)
                df['1st-Order Category'] = category_list_1
                first0_summary = df.groupby(
                    '1st-Order Category')['Avg. Score (%)'].mean().reset_index()

                # Start from A22 and find the last filled cell in column A
                current_row = 22
                while ws[f"A{current_row}"].value is not None:  # Check if cell is filled
                    current_row += 1  # Move down

                # Move one row below the last filled cell
                start_row = current_row + 1

                # Insert 12 blank rows after the last row written
                ws.insert_rows(current_row, 12)

                # Add the overall average
                ws["A105"] = "Overall Average"
                ws["B105"] = int(overall_avg)

                # use a custom sort for the summary table
                first0_summary['1st-Order Category'] = pd.Categorical(
                    first0_summary['1st-Order Category'], [
                        "THRIVE", "Just Leader",
                    ])
                first0_summary = first0_summary.sort_values(
                    "1st-Order Category")

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                first0_new = first0_summary.set_index("1st-Order Category").T

                # header column
                ws["A107"] = "1st-Order Category"
                ws["A108"] = "Avg. Score (%)"

                # copy in heading labels
                ws["B107"] = "THRIVE"
                ws["C107"] = "Just Leader"

                row = 108  # Starting point
                for _, row_data in first0_new.iterrows():
                    ws[f"B{row}"] = row_data["THRIVE"]
                    ws[f"C{row}"] = row_data["Just Leader"]
                    row += 1  # Move to the next row

                # Second level breakdown
                repetitions_2 = [10, 10, 10, 10, 10, 10, 5, 5, 5, 5]
                category_list_2 = []
                for category, rep in zip(category_2, repetitions_2):
                    category_list_2.extend([category] * rep)
                df['2nd-Order Category'] = category_list_2
                second0_summary = df.groupby(
                    ['1st-Order Category', '2nd-Order Category'])['Avg. Score (%)'].mean().reset_index().drop(columns='1st-Order Category')

                # use a custom sort for the summary table
                second0_summary['2nd-Order Category'] = pd.Categorical(
                    second0_summary['2nd-Order Category'], [
                        "Trust", "Health", "Relationships", "Impact", "Value", "Engagement", "See the Whole Playing Field", "Build Cultural Competency", "Give Power Away", "Take Bold, Courageous Action"
                    ])
                second0_summary = second0_summary.sort_values(
                    "2nd-Order Category")

                # header rows
                ws["A110"] = "2nd-Order Category"
                ws["A111"] = "Avg. Score (%)"

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                second0_new = second0_summary.set_index("2nd-Order Category").T

                # New code for wide format
                row = 111  # Starting point
                for _, row_data in second0_new.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Health"]
                    ws[f"D{row}"] = row_data["Relationships"]
                    ws[f"E{row}"] = row_data["Impact"]
                    ws[f"F{row}"] = row_data["Value"]
                    ws[f"G{row}"] = row_data["Engagement"]
                    ws[f"H{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"I{row}"] = row_data["Build Cultural Competency"]
                    ws[f"J{row}"] = row_data["Give Power Away"]
                    ws[f"K{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                # Copy in 2nd-Order Category labels
                ws["B110"] = "Trust"
                ws["C110"] = "Health"
                ws["D110"] = "Relationships"
                ws["E110"] = "Impact"
                ws["F110"] = "Value"
                ws["G110"] = "Engagement"
                ws["H110"] = "See the Whole Playing Field"
                ws["I110"] = "Build Cultural Competency"
                ws["J110"] = "Give Power Away"
                ws["K110"] = "Take Bold, Courageous Action"

                # Third level breakdown
                repetitions_3 = [5, 5, 5, 5, 5, 5, 5, 5,
                                 5, 5, 5, 5, 2, 3, 2, 3, 2, 3, 2, 3]
                category_list_3 = []
                for category, rep in zip(category_3, repetitions_3):
                    category_list_3.extend([category] * rep)
                df['3rd-Order Category'] = category_list_3

                # create 3rd-order contat column
                third0_summary = df[df['2nd-Order Category'] != 'Health'].groupby(
                    ['2nd-Order Category', '3rd-Order Category'])['Avg. Score (%)'].mean().reset_index()

                # header rows
                ws["A113"] = "3rd-Order Category"
                ws["A114"] = "Leader Avg. Score (%)"
                ws["A115"] = "Team Avg. Score (%)"

                # Per Uncle David's request, show 3rd-order category averages in wide format instead of long format
                third0_leader = third0_summary[third0_summary['3rd-Order Category']
                                               == 'Leader'].drop(columns='3rd-Order Category')
                third0_leader = third0_leader.set_index("2nd-Order Category").T

                # New code for wide format
                row = 114  # Starting point
                for _, row_data in third0_leader.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Relationships"]
                    ws[f"D{row}"] = row_data["Impact"]
                    ws[f"E{row}"] = row_data["Value"]
                    ws[f"F{row}"] = row_data["Engagement"]
                    ws[f"G{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"H{row}"] = row_data["Build Cultural Competency"]
                    ws[f"I{row}"] = row_data["Give Power Away"]
                    ws[f"J{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                third0_team = third0_summary[third0_summary['3rd-Order Category']
                                             == 'Team'].drop(columns='3rd-Order Category')
                third0_team = third0_team.set_index("2nd-Order Category").T

                row = 115  # Starting point
                for _, row_data in third0_team.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Relationships"]
                    ws[f"D{row}"] = row_data["Impact"]
                    ws[f"E{row}"] = row_data["Value"]
                    ws[f"F{row}"] = row_data["Engagement"]
                    ws[f"G{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"H{row}"] = row_data["Build Cultural Competency"]
                    ws[f"I{row}"] = row_data["Give Power Away"]
                    ws[f"J{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                # Copy in 3rd-Order Category labels
                ws["B113"] = "Trust"
                ws["C113"] = "Relationships"
                ws["D113"] = "Impact"
                ws["E113"] = "Value"
                ws["F113"] = "Engagement"
                ws["G113"] = "See the Whole Playing Field"
                ws["H113"] = "Build Cultural Competency"
                ws["I113"] = "Give Power Away"
                ws["J113"] = "Take Bold, Courageous Action"

                # Define the range of cells
                start_row = 24
                end_row = 116
                start_col = 1  # Column A (1-indexed)
                end_col = 11  # Column C (1-indexed)

                # Define the font style
                custom_font = Font(name="Arial", size=11)

                # Apply the font style to each cell in the range
                for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.font = custom_font

                # format the headers
                ws["A105"].font = Font(name="Arial", size=12, bold=True)
                ws["B105"].font = Font(name="Arial", size=12, bold=True)

                ws["A107"].font = Font(name="Arial", size=11, bold=True)
                ws["A108"].font = Font(name="Arial", size=11, bold=True)

                ws["A110"].font = Font(name="Arial", size=11, bold=True)
                ws["A111"].font = Font(name="Arial", size=11, bold=True)

                ws["A113"].font = Font(name="Arial", size=11, bold=True)
                ws["A114"].font = Font(name="Arial", size=11, bold=True)
                ws["A115"].font = Font(name="Arial", size=11, bold=True)

                ws["A107"].alignment = copy(ws["C23"].alignment)
                ws["A108"].alignment = copy(ws["C23"].alignment)

                ws["A110"].alignment = copy(ws["C23"].alignment)
                ws["A111"].alignment = copy(ws["C23"].alignment)
                ws["A113"].alignment = copy(ws["C23"].alignment)
                ws["A114"].alignment = copy(ws["C23"].alignment)
                ws["A115"].alignment = copy(ws["C23"].alignment)

                # rounding
                for row in ws.iter_rows(min_row=105, max_row=116, min_col=2, max_col=11):
                    for cell in row:
                        # Check if cell contains a numeric value
                        if isinstance(cell.value, (int, float)):

                            cell.number_format = '0'

                # row height
                for row in range(104, 116):
                    ws.row_dimensions[row].height = 15

            else:  # Team template

                template = 'Team'
                templates_list.append(template)
                df_team = df.copy()

                # First level breakdown
                repetitions_1 = [60, 20]
                category_list_1 = []
                for category, rep in zip(category_1, repetitions_1):
                    category_list_1.extend([category] * rep)
                df['1st-Order Category'] = category_list_1
                first0_summary = df.groupby(
                    '1st-Order Category')['Avg. Score (%)'].mean().reset_index()

                # Start from A22 and find the last filled cell in column A
                current_row = 22
                while ws[f"A{current_row}"].value is not None:  # Check if cell is filled
                    current_row += 1  # Move down

                # Move one row below the last filled cell
                start_row = current_row + 1

                # Insert blank rows after the last row written
                ws.insert_rows(current_row, 12)

                # Add the overall average
                ws["A105"] = "Overall Average"
                ws["B105"] = overall_avg

                # use a custom sort for the summary table
                first0_summary['1st-Order Category'] = pd.Categorical(
                    first0_summary['1st-Order Category'], [
                        "THRIVE", "Just Leader",
                    ])
                first0_summary = first0_summary.sort_values(
                    "1st-Order Category")

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                first0_new = first0_summary.set_index("1st-Order Category").T

                # header column
                ws["A107"] = "1st-Order Category"
                ws["A108"] = "Avg. Score (%)"

                # copy in heading labels
                ws["B107"] = "THRIVE"
                ws["C107"] = "Just Leader"

                row = 108  # Starting point
                for _, row_data in first0_new.iterrows():
                    ws[f"B{row}"] = row_data["THRIVE"]
                    ws[f"C{row}"] = row_data["Just Leader"]
                    row += 1  # Move to the next row

                # Second level breakdown
                repetitions_2 = [10, 10, 10, 10, 10, 10, 5, 5, 5, 5]
                category_list_2 = []
                for category, rep in zip(category_2, repetitions_2):
                    category_list_2.extend([category] * rep)
                df['2nd-Order Category'] = category_list_2
                second0_summary = df.groupby(
                    ['1st-Order Category', '2nd-Order Category'])['Avg. Score (%)'].mean().reset_index().drop(columns='1st-Order Category')

                # use a custom sort for the summary table
                second0_summary['2nd-Order Category'] = pd.Categorical(
                    second0_summary['2nd-Order Category'], [
                        "Trust", "Health", "Relationships", "Impact", "Value", "Engagement", "See the Whole Playing Field", "Build Cultural Competency", "Give Power Away", "Take Bold, Courageous Action"
                    ])
                second0_summary = second0_summary.sort_values(
                    "2nd-Order Category")

                # header rows
                ws["A110"] = "2nd-Order Category"
                ws["A111"] = "Avg. Score (%)"

                # Per Uncle David's request, show 2nd-order category averages in wide format instead of long format
                second0_new = second0_summary.set_index("2nd-Order Category").T

                # New code for wide format
                row = 111  # Starting point
                for _, row_data in second0_new.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Health"]
                    ws[f"D{row}"] = row_data["Relationships"]
                    ws[f"E{row}"] = row_data["Impact"]
                    ws[f"F{row}"] = row_data["Value"]
                    ws[f"G{row}"] = row_data["Engagement"]
                    ws[f"H{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"I{row}"] = row_data["Build Cultural Competency"]
                    ws[f"J{row}"] = row_data["Give Power Away"]
                    ws[f"K{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                # Copy in 2nd-Order Category labels
                ws["B110"] = "Trust"
                ws["C110"] = "Health"
                ws["D110"] = "Relationships"
                ws["E110"] = "Impact"
                ws["F110"] = "Value"
                ws["G110"] = "Engagement"
                ws["H110"] = "See the Whole Playing Field"
                ws["I110"] = "Build Cultural Competency"
                ws["J110"] = "Give Power Away"
                ws["K110"] = "Take Bold, Courageous Action"

                # Third level breakdown
                repetitions_3 = [5, 5, 5, 5, 5, 5, 5, 5,
                                 5, 5, 5, 5, 2, 3, 2, 3, 2, 3, 2, 3]
                category_list_3 = []
                for category, rep in zip(category_3, repetitions_3):
                    category_list_3.extend([category] * rep)
                df['3rd-Order Category'] = category_list_3

                # create 3rd-order contat column
                third0_summary = df[df['2nd-Order Category'] != 'Health'].groupby(
                    ['2nd-Order Category', '3rd-Order Category'])['Avg. Score (%)'].mean().reset_index()

                # header rows
                ws["A113"] = "3rd-Order Category"
                ws["A114"] = "Leader Avg. Score (%)"
                ws["A115"] = "Team Avg. Score (%)"

                # Per Uncle David's request, show 3rd-order category averages in wide format instead of long format
                third0_leader = third0_summary[third0_summary['3rd-Order Category']
                                               == 'Leader'].drop(columns='3rd-Order Category')
                third0_leader = third0_leader.set_index("2nd-Order Category").T

                # New code for wide format
                row = 114  # Starting point
                for _, row_data in third0_leader.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Relationships"]
                    ws[f"D{row}"] = row_data["Impact"]
                    ws[f"E{row}"] = row_data["Value"]
                    ws[f"F{row}"] = row_data["Engagement"]
                    ws[f"G{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"H{row}"] = row_data["Build Cultural Competency"]
                    ws[f"I{row}"] = row_data["Give Power Away"]
                    ws[f"J{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                third0_team = third0_summary[third0_summary['3rd-Order Category']
                                             == 'Team'].drop(columns='3rd-Order Category')
                third0_team = third0_team.set_index("2nd-Order Category").T

                row = 115  # Starting point
                for _, row_data in third0_team.iterrows():
                    ws[f"B{row}"] = row_data["Trust"]
                    ws[f"C{row}"] = row_data["Relationships"]
                    ws[f"D{row}"] = row_data["Impact"]
                    ws[f"E{row}"] = row_data["Value"]
                    ws[f"F{row}"] = row_data["Engagement"]
                    ws[f"G{row}"] = row_data["See the Whole Playing Field"]
                    ws[f"H{row}"] = row_data["Build Cultural Competency"]
                    ws[f"I{row}"] = row_data["Give Power Away"]
                    ws[f"J{row}"] = row_data["Take Bold, Courageous Action"]
                    row += 1

                # Copy in 3rd-Order Category labels
                ws["B113"] = "Trust"
                ws["C113"] = "Relationships"
                ws["D113"] = "Impact"
                ws["E113"] = "Value"
                ws["F113"] = "Engagement"
                ws["G113"] = "See the Whole Playing Field"
                ws["H113"] = "Build Cultural Competency"
                ws["I113"] = "Give Power Away"
                ws["J113"] = "Take Bold, Courageous Action"

                # Define the range of cells
                start_row = 24
                end_row = 116
                start_col = 1  # Column A (1-indexed)
                end_col = 11  # Column C (1-indexed)

                # Define the font style
                custom_font = Font(name="Arial", size=11)

                # Apply the font style to each cell in the range
                for row in ws.iter_rows(min_row=start_row, max_row=end_row, min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.font = custom_font

                # format the headers
                ws["A105"].font = Font(name="Arial", size=12, bold=True)
                ws["B105"].font = Font(name="Arial", size=12, bold=True)

                ws["A107"].font = Font(name="Arial", size=11, bold=True)
                ws["A108"].font = Font(name="Arial", size=11, bold=True)

                ws["A110"].font = Font(name="Arial", size=11, bold=True)
                ws["A111"].font = Font(name="Arial", size=11, bold=True)

                ws["A113"].font = Font(name="Arial", size=11, bold=True)
                ws["A114"].font = Font(name="Arial", size=11, bold=True)
                ws["A115"].font = Font(name="Arial", size=11, bold=True)

                ws["A107"].alignment = copy(ws["C23"].alignment)
                ws["A108"].alignment = copy(ws["C23"].alignment)

                ws["A110"].alignment = copy(ws["C23"].alignment)
                ws["A111"].alignment = copy(ws["C23"].alignment)
                ws["A113"].alignment = copy(ws["C23"].alignment)
                ws["A114"].alignment = copy(ws["C23"].alignment)
                ws["A115"].alignment = copy(ws["C23"].alignment)

                # rounding
                for row in ws.iter_rows(min_row=105, max_row=117, min_col=2, max_col=12):
                    for cell in row:
                        # Check if cell contains a numeric value
                        if isinstance(cell.value, (int, float)):
                            # Round to 1 decimal place
                            cell.number_format = '0'

                # row height
                for row in range(104, 116):
                    ws.row_dimensions[row].height = 15

                # get standard deviation values in template
                std_dev_values = []

                # Start iterating from the first cell in column C
                column = 3  # Column C
                row = 1  # Start at the first row
                max_row = 5000

                # Iterate through column C until the end
                while row <= max_row:
                    cell_value = ws.cell(row=row, column=column).value

                    # Check if the cell contains 'Standard Deviation'
                    if cell_value == "Standard Deviation":
                        # Get the value in the cell below
                        next_cell_value = ws.cell(
                            row=row + 1, column=column).value
                        # if next_cell_value is not None, append to std_dev_values as a float
                        if next_cell_value is not None:
                            next_cell_value = float(next_cell_value)
                            std_dev_values.append(next_cell_value)

                        # Move the row pointer down by 2 to skip the value we just processed
                        row += 2
                    else:
                        # Move to the next row
                        row += 1

            # For the Leader and Team templates, there will be a 3rd order category
            if template == 'Leader' or template == 'Team':

                # label Column G, which won't be necessary for the other 2 templates
                ws["G23"] = "3rd-Order Category"
                ws["G23"].font = copy(ws["C23"].font)
                ws["G23"].fill = copy(ws["C23"].fill)
                ws["G23"].alignment = copy(ws["C23"].alignment)

                # remove 3rd-order category when 2nd-order == 'Health'
                df["3rd-Order Category"] = np.where(
                    df["2nd-Order Category"] == 'Health', 'n/a', df["3rd-Order Category"])

                row = 24  # Starting from row 24
                for idx, row_data in df.iterrows():
                    ws[f"A{row}"] = row_data["Questions"]
                    ws[f"B{row}"] = row_data["Difficulty"]
                    ws[f"C{row}"] = row_data["Avg. Score (%)"]
                    ws[f"D{row}"] = row_data["Question Order"]
                    ws[f"E{row}"] = row_data["1st-Order Category"]
                    ws[f"F{row}"] = row_data["2nd-Order Category"]
                    ws[f"G{row}"] = row_data["3rd-Order Category"]
                    row += 1  # Move to the next row
                ws.column_dimensions['G'].width = 18

            # remove 3rd-order category for the 'Leader' template
            if template == 'Leader':
                ws.delete_rows(113, 4)

            if template == 'Team':
                start_col, end_col = 3, 10

                # Loop backward through columns to shift values right
                for col in range(end_col, start_col - 1, -1):  # From J to C
                    for row in range(113, 116):  # Rows 113 to 115
                        ws.cell(row=row, column=col + 1,
                                value=ws.cell(row=row, column=col).value)
                        # Clear old cell
                        ws.cell(row=row, column=col, value=None)

                # now insert 'Health' into column C
                ws["C113"] = "Health"
                ws["C114"] = "N/A"
                ws["C115"] = "N/A"

                # now right-align cells C114 and C115
                ws["C114"].alignment = Alignment(horizontal="right")
                ws["C115"].alignment = Alignment(horizontal="right")

            # For the Review and No Leader templates, there will be no 3rd order category
            else:
                row = 24  # Starting from row 24
                for idx, row_data in df.iterrows():
                    ws[f"A{row}"] = row_data["Questions"]
                    ws[f"B{row}"] = row_data["Difficulty"]
                    ws[f"C{row}"] = row_data["Avg. Score (%)"]
                    ws[f"D{row}"] = row_data["Question Order"]
                    ws[f"E{row}"] = row_data["1st-Order Category"]
                    ws[f"F{row}"] = row_data["2nd-Order Category"]
                    row += 1  # Move to the next row

            # for only the Team, Review, and No Leader templates, shift data over to make room for the standard deviation values
            if template != 'Leader':
                # Find the data range starting from D23
                start_row = 23
                start_col = 4  # Column D

                # Find the last row in column D (stop when an empty cell is encountered)
                end_row = start_row
                while ws.cell(row=end_row, column=start_col).value is not None:
                    end_row += 1
                end_row -= 1  # Adjust to the last filled row

                # Find the last column (stop when an empty cell is encountered in the header row)
                end_col = start_col
                while ws.cell(row=start_row, column=end_col).value is not None:
                    end_col += 1
                end_col -= 1  # Adjust to the last filled column

                # Shift data one column to the right
                for row in range(start_row, end_row + 1):
                    # Move backward to avoid overwriting
                    for col in range(end_col, start_col - 1, -1):
                        source_cell = ws.cell(row=row, column=col)
                        target_cell = ws.cell(row=row, column=col + 1)

                        # Copy value
                        target_cell.value = source_cell.value

                        # Copy style if present
                        if source_cell.has_style:
                            target_cell._style = source_cell._style

                        # Clear the original cell
                        source_cell.value = None

                # Now input the standard deviation values
                ws["D23"] = "Standard deviation"
                start_row = 24
                start_column = 4  # Column D

                # Write the values from the list into column D, starting at D24
                for i, value in enumerate(std_dev_values):
                    cell = ws.cell(row=start_row + i, column=start_column)
                    cell.value = float(value)

            # any final adjustments to the table
            ws["C23"] = "Avg. Score (%)"
            ws.column_dimensions['A'].width = 50
            ws.column_dimensions['B'].width = 14
            ws.column_dimensions['C'].width = 17
            ws.column_dimensions['D'].width = 17
            ws.column_dimensions['E'].width = 14
            ws.column_dimensions['F'].width = 19
            ws.column_dimensions['G'].width = 19
            ws.column_dimensions['H'].width = 23
            ws.column_dimensions['I'].width = 22
            ws.column_dimensions['J'].width = 15
            ws.column_dimensions['K'].width = 21

            # Add "_clean" suffix to the file name before the extension
            clean_file_name = f"{uploaded_file.name.rsplit('.', 1)[0]}_clean.xlsx"

            # Save the modified workbook to a BytesIO object
            cleaned_file = io.BytesIO()
            wb.save(cleaned_file)
            cleaned_file.seek(0)

            # Store the new name and file data
            files_to_download.append((clean_file_name, cleaned_file))

        # If applicable, create the comparison dataframe between Team and Leader
        if 'Leader' in templates_list and 'Team' in templates_list:
            # create the comparison dataframe
            df_comparison = pd.merge(
                df_leader, df_team, on="Question Order", suffixes=("_Leader", "_Team")).drop(columns=["Difficulty_Leader", "Difficulty_Team"])
            df_comparison = df_comparison.rename(
                columns={"Avg. Score (%)_Leader": "Score_Leader", "Avg. Score (%)_Team": "Score_Team"})
            df_comparison = df_comparison[[
                'Question Order', 'Questions_Leader', 'Score_Leader', 'Questions_Team', 'Score_Team']]
            df_comparison['Score_delta'] = df_comparison['Score_Leader'] - \
                df_comparison['Score_Team']
            df_comparison = df_comparison.sort_values(
                by="Score_delta", ascending=False)

            df_comparison_name = "Leader-Team_Comparison.xlsx"

            # Save the dataframe to a BytesIO object in Excel format
            comparison_file = io.BytesIO()
            with pd.ExcelWriter(comparison_file, engine='openpyxl') as writer:
                df_comparison.to_excel(
                    writer, index=False, sheet_name='Comparison')

                # Access the worksheet to set column widths
                worksheet = writer.sheets['Comparison']  # Get the worksheet
                worksheet.column_dimensions['A'].width = 13
                worksheet.column_dimensions['B'].width = 75
                worksheet.column_dimensions['C'].width = 13
                worksheet.column_dimensions['D'].width = 75
                worksheet.column_dimensions['E'].width = 13
                worksheet.column_dimensions['F'].width = 13

            # Move to the beginning of the BytesIO buffer
            comparison_file.seek(0)

            # Append the file to the files_to_download list
            files_to_download.append((df_comparison_name, comparison_file))
            st.markdown(f'''
                <p style="font-size: 18px; font-weight: 100; text-align: center; margin-top: 0px; margin-bottom: 40px; color: #fefefe;">
                    <i><b>Note:</b> You have uploaded a Leader and a Team template. You will find a comparison file with the zipped bundle in your Downloads folder when you press the button below.</i>
                </p>
            ''', unsafe_allow_html=True)

        # Handle downloading the files
        if len(files_to_download) > 1:  # Create a ZIP file for multiple uploads
            zip_buffer = io.BytesIO()
            with zipfile.ZipFile(zip_buffer, "w") as zip_file:
                for file_name, file_data in files_to_download:
                    zip_file.writestr(file_name, file_data.getvalue())
            zip_buffer.seek(0)

            # Provide download button for the ZIP file
            st.download_button(
                label=f"Clean & Download Files",
                data=zip_buffer,
                file_name="uploaded_files_clean.zip",
                mime="application/zip"
            )

        else:  # Provide download button for a single file
            file_name, file_data = files_to_download[0]
            st.download_button(
                label="Clean & Download File",
                data=file_data,
                file_name=file_name,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )


# Run the app
if __name__ == "__main__":
    main()
